# -*- coding: utf-8 -*-
"""
@author: ashwinkumars
"""

import time
import logging
from openpyxl import load_workbook

from ABAInputXMLDataProcessor import load_training_data
from ABADocumentCreater import create_document_for_Query
from ABAUtils import sanitize_doc
from ABAUtils import get_assigned_dev
from ABAUtils import get_similarity_vector

logger = logging.getLogger(__name__)


def get_indices(worksheet, docIdColName) :    
    retRowIndex = 0
    retColumnIndex = 0
    
    for rowIndex in range(worksheet.max_row) :
        for columnIndex in range(worksheet.max_column) :
            if docIdColName == worksheet.cell(row = rowIndex + 1, column = columnIndex + 1).value :
                retRowIndex = rowIndex + 1
                retColumnIndex = columnIndex + 1
                break
        
    return retRowIndex, retColumnIndex

def get_dev_value(queryIdList, assignedDevList, queryId) :
    devVal = ''
    
    for qId, dVal in zip(queryIdList, assignedDevList) :
        if qId == queryId :
            devVal = dVal
    
    return devVal

def write_query_data(dataABA, queryIdList, assignedDevList):
    workbook = load_workbook(dataABA.docPath)
    worksheet = workbook[dataABA.sheetName]    
    
    queryIdRowIndex, queryIdColumnIndex = get_indices(worksheet, dataABA.docIdColName)
    devColNameRowIndex, devColNameColumnIndex = get_indices(worksheet, dataABA.devColName)
    
    #queryIdRowIndex will be same as devColNameRowIndex
    if queryIdRowIndex != 0 and queryIdRowIndex < worksheet.max_row :
        for rowIndex in range(queryIdRowIndex, worksheet.max_row) :
            queryId = worksheet.cell(rowIndex + 1, queryIdColumnIndex).value
            devVal = get_dev_value(queryIdList, assignedDevList, queryId)
            
            if devVal != '' :
                worksheet.cell(rowIndex + 1, devColNameColumnIndex).value = devVal
            
    
# =============================================================================
#     for index, row in enumerate(worksheet.iter_rows()) :
#         for cell in row :
#             print(index)
#             print(worksheet.cell(row=1, column=index+1).value)
#             print(cell.value)
# =============================================================================
    #worksheet.cell(row=4, column=5)     
    
    workbook.save(dataABA.docPath)


def run_query_ABA(queryIdList, queryList, trainInputFile) :    
    assignedDevList = []
    devMap, idf, tfidfRepresentation = load_training_data(trainInputFile)    
      
    for queryId, query in zip(queryIdList, queryList) :
        similarityVector = get_similarity_vector(tfidfRepresentation, idf, query)
        assignedDev = get_assigned_dev(similarityVector, devMap)  
        assignedDevList.append(assignedDev)
        
        logger.info("Id = " + str(queryId) + ", Assigned developer = " + assignedDev)
        
    return queryIdList, assignedDevList

def execute_query(dataABA) :    
    
    sanitizedQueryIdList = []
    sanitizedQueryList = []
    
    start = time.time()
    logger.info('Query started for file : ' + dataABA.docPath + '.')
    
    queryIdList, queryList = create_document_for_Query(dataABA)

    #remove empty dev with no docs
    for queryId, queryList in zip(queryIdList, queryList) :
        queryList = sanitize_doc(queryList)
        if queryList != '' :
            sanitizedQueryIdList.append(queryId)
            sanitizedQueryList.append(queryList)
    
    if len(sanitizedQueryList) > 0 :
        queryIdList, assignedDevList = run_query_ABA(sanitizedQueryIdList, sanitizedQueryList, dataABA.trainFilePath)
        write_query_data(dataABA, queryIdList, assignedDevList)        
    
    end = time.time()
    logger.info("Total time :" + str(end - start) + '\n')
    logger.info('Query complete for ' + dataABA.docPath + '.')
